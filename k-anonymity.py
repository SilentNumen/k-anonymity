import openpyxl
import pandas as pd


def get_my_table(file):
    user = []
    book = openpyxl.load_workbook(file)
    sheet1 = book['Sheet1']
    len_row = sheet1.max_row  # 列
    len_column = sheet1.max_column  # 行
    for d in range(1, len_row + 1):
        item = []
        for j in range(1, len_column + 1):
            item.append(sheet1.cell(d, j).value)
        user.append(item)
    del user[0]
    variables0 = ['user_name', 'user_age', 'user_wage', 'user_telephone', 'user_illness']
    df = pd.DataFrame(user, columns=variables0)
    print('原始数据表格')
    for f in range(64):
        print('-', end='')
    print()
    print(df)
    print()
    return user


def similarity_recongnition(n1, n2):
    m1 = ''
    for c in range(len(n1)):
        if n1[c] == n2[c]:
            m1 = m1 + n1[c]
        if n1[c] != n2[c]:
            break
    for c in range(len(n1) - len(m1)):
        m1 = m1 + '*'
    return m1


class k_anonymity:
    def __init__(self, table0):
        self.table = table0
        self.len = len(table0)
        self.index_buttle = []

    def int_min(self, list0, m):
        weight = []
        for b in range(len(self.table)):
            weight.append(0)

        user_age = []
        for b in self.table:
            user_age.append(b)
        while len(user_age) > m:
            index = 0
            max_age = 0
            for b in range(len(user_age)):
                if (int(user_age[b][2]) - int(list0[2])) ** 2 > max_age ** 2:
                    max_age = int(user_age[b][2]) - int(list0[2])
                    index = b
            del user_age[index]
        for b in user_age:
            for j in range(self.len):
                if b == self.table[j]:
                    weight[j] = weight[j] + 1

        user_wage = []
        for b in self.table:
            user_wage.append(b)
        while len(user_wage) > m:
            index = 0
            max_wage = 0
            for b in range(len(user_wage)):
                if (int(user_wage[b][2]) - int(list0[2])) ** 2 > max_wage ** 2:
                    max_wage = int(user_wage[b][2]) - int(list0[2])
                    index = b
            del user_wage[index]
        for b in user_wage:
            for j in range(self.len):
                if b == self.table[j]:
                    weight[j] = weight[j] + 1

        user_telephone = []
        for b in self.table:
            user_telephone.append(b)
        while len(user_telephone) > m:
            index = 0
            max_telephone = 0
            for b in range(len(user_telephone)):
                if (int(user_telephone[b][2]) - int(list0[2])) ** 2 > max_telephone ** 2:
                    max_telephone = int(user_telephone[b][2]) - int(list0[2])
                    index = b
            del user_telephone[index]
        for b in user_telephone:
            for j in range(self.len):
                if b == self.table[j]:
                    weight[j] = weight[j] + 1

        group = []
        count = 0
        while count != m:
            index = 0
            max_weight = 0
            for b in range(len(weight)):
                if weight[b] > max_weight:
                    max_weight = weight[b]
                    index = b
            weight[index] = 0
            count = count + 1
            self.index_buttle.append(index)
            group.append(self.table[index])
        return group

    def anonymity(self, n):
        work = []
        for a in range(self.len):
            if a not in self.index_buttle:
                my_group = self.int_min(self.table[a], n)
                for item in my_group:
                    work.append(item)
        new_k_table = []
        for e in range(0, len(work), n):
            k_table = []
            for j in range(n):
                k_table.append(work[e + j])
            for item0 in k_table:
                str0 = ''
                str1 = ''
                str2 = ''
                for item1 in k_table:
                    if item1 != item0:
                        str0 = similarity_recongnition(str(item0[1]), str(item1[1]))
                        str1 = similarity_recongnition(str(item0[2]), str(item1[2]))
                        str2 = similarity_recongnition(str(item0[3]), str(item1[3]))
                        item1[1] = str0
                        item1[2] = str1
                        item1[3] = str2
                item0[1] = str0
                item0[2] = str1
                item0[3] = str2
            for T in k_table:
                new_k_table.append(T)
        return new_k_table


if __name__ == '__main__':
    k = 3
    file_path = 'user/user2.xlsx'
    user_table = get_my_table(file_path)
    my_table = k_anonymity(user_table)
    table = my_table.anonymity(k)
    print(str(k) + '-匿名后数据')
    for i in range(64):
        print('-', end='')
    print()
    variables = ['user_name', 'user_age', 'user_wage', 'user_telephone', 'user_illness']
    dt = pd.DataFrame(table, columns=variables)
    print(dt)
