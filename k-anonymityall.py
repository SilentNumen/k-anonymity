import openpyxl
import pandas as pd
import numpy as np


def get_noisy_digit(value0, epsilon0):
    u = np.random.random() - 0.5
    noisy_digit = 0 - value0 / epsilon0 * np.sign(u) * np.log(1.0 - 2 * np.abs(u))
    return np.rint(noisy_digit)


def get_my_table(file):
    user = []
    book = openpyxl.load_workbook(file)
    sheet1 = book['Sheet1']
    len_row = sheet1.max_row  # 列
    len_column = sheet1.max_column  # 行
    for d in range(1, len_row + 1):
        item = []
        for j in range(1, len_column + 1):
            item.append(sheet1.cell(d, j).value)
        user.append(item)
    del user[0]
    variables0 = ['user_name', 'user_age', 'user_wage', 'user_telephone', 'user_illness']
    df = pd.DataFrame(user, columns=variables0)
    print('原始数据表格')
    for f in range(64):
        print('-', end='')
    print()
    print(df)
    print()
    return user


class k_anonymity:
    def __init__(self, table0):
        self.table = table0
        self.len = len(table0)
        self.index_buttle = []

    def int_min(self, list0, m):
        weight = []
        for b in range(len(self.table)):
            weight.append(0)

        user_age = []
        for b in self.table:
            user_age.append(b)
        while len(user_age) > m:
            index = 0
            max_age = 0
            for b in range(len(user_age)):
                if (int(user_age[b][2]) - int(list0[2])) ** 2 > max_age ** 2:
                    max_age = int(user_age[b][2]) - int(list0[2])
                    index = b
            del user_age[index]
        for b in user_age:
            for j in range(self.len):
                if b == self.table[j]:
                    weight[j] = weight[j] + 1

        user_wage = []
        for b in self.table:
            user_wage.append(b)
        while len(user_wage) > m:
            index = 0
            max_wage = 0
            for b in range(len(user_wage)):
                if (int(user_wage[b][2]) - int(list0[2])) ** 2 > max_wage ** 2:
                    max_wage = int(user_wage[b][2]) - int(list0[2])
                    index = b
            del user_wage[index]
        for b in user_wage:
            for j in range(self.len):
                if b == self.table[j]:
                    weight[j] = weight[j] + 1

        user_telephone = []
        for b in self.table:
            user_telephone.append(b)
        while len(user_telephone) > m:
            index = 0
            max_telephone = 0
            for b in range(len(user_telephone)):
                if (int(user_telephone[b][2]) - int(list0[2])) ** 2 > max_telephone ** 2:
                    max_telephone = int(user_telephone[b][2]) - int(list0[2])
                    index = b
            del user_telephone[index]
        for b in user_telephone:
            for j in range(self.len):
                if b == self.table[j]:
                    weight[j] = weight[j] + 1

        group = []
        count = 0
        while count != m:
            index = 0
            max_weight = 0
            for b in range(len(weight)):
                if weight[b] > max_weight:
                    max_weight = weight[b]
                    index = b
            weight[index] = 0
            count = count + 1
            self.index_buttle.append(index)
            group.append(self.table[index])
        return group

    def anonymity(self, n):
        work = []
        for a in range(self.len):
            if a not in self.index_buttle:
                my_group = self.int_min(self.table[a], n)
                for item in my_group:
                    work.append(item)

        new_k_table = []
        for e in range(0, len(work), n):
            k_table = []
            for j in range(n):
                k_table.append(work[e + j])

            # 差分隐私过程
            value = 1000
            epsilon = 3
            print('加噪前：', end='')
            print(k_table)
            for ou in range(len(k_table)):
                k_table[ou][2] = int(k_table[ou][2])
                k_table[ou][2] = int((k_table[ou][2] + get_noisy_digit(value, epsilon)))
            print('加噪后：', end='')
            print(k_table)
            print()

            for item0 in k_table:
                m = ''
                item0[1] = str(item0[1])
                for number1 in range(len(item0[1]) - 1):
                    m = m + item0[1][number1]
                for number2 in range(len(item0[1]) - 1):
                    m = m + '*'
                item0[1] = m

                m = ''
                item0[2] = str(item0[2])
                for number1 in range(len(item0[2]) - 3):
                    m = m + item0[2][number1]
                for number2 in range(len(item0[2]) - 1):
                    m = m + '*'
                item0[2] = m

                m = ''
                item0[3] = str(item0[3])
                for number1 in range(len(item0[3]) - 8):
                    m = m + item0[3][number1]
                for number2 in range(len(item0[3]) - 3):
                    m = m + '*'
                item0[3] = m
            for T in k_table:
                new_k_table.append(T)

        return new_k_table


if __name__ == '__main__':
    k = 3
    file_path = 'user/user2.xlsx'
    user_table = get_my_table(file_path)
    my_table = k_anonymity(user_table)
    table = my_table.anonymity(k)
    print(str(k) + '-匿名后数据')
    for i in range(64):
        print('-', end='')
    print()
    variables = ['user_name', 'user_age', 'user_wage', 'user_telephone', 'user_illness']
    dt = pd.DataFrame(table, columns=variables)
    print(dt)
